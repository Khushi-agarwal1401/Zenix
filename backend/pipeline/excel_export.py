"""
Excel Export Module for Zenix AI.
Provides export functionality to Excel format.
"""

import os
import logging
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export data to Excel format.
    Uses openpyxl if available, falls back to CSV.
    """

    def __init__(self):
        self._openpyxl_available = None

    def _check_openpyxl(self) -> bool:
        """Check if openpyxl is available."""
        if self._openpyxl_available is not None:
            return self._openpyxl_available
        try:
            import openpyxl
            self._openpyxl_available = True
        except ImportError:
            self._openpyxl_available = False
        return self._openpyxl_available

    def export_to_excel(self, data: List[Dict[str, Any]], output_path: str,
                       sheet_name: str = "Data") -> bool:
        """
        Export data to Excel file.

        Args:
            data: List of dictionaries to export
            output_path: Path to save the Excel file
            sheet_name: Name of the worksheet

        Returns:
            True if successful, False otherwise
        """
        if not self._check_openpyxl():
            logger.warning("openpyxl not installed. Falling back to CSV export.")
            return self.export_to_csv(data, output_path.replace('.xlsx', '.csv'))

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            if not data:
                return False

            # Get headers from first item
            headers = list(data[0].keys())

            # Style for header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Write data rows
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = max(
                    len(str(ws.cell(row=r, column=col).value or ""))
                    for r in range(1, len(data) + 2)
                )
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_length + 2, 50)

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Save file
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)

            logger.info(f"Excel file exported: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return self.export_to_csv(data, output_path.replace('.xlsx', '.csv'))

    def export_to_csv(self, data: List[Dict[str, Any]], output_path: str) -> bool:
        """
        Export data to CSV file (fallback).

        Args:
            data: List of dictionaries to export
            output_path: Path to save the CSV file

        Returns:
            True if successful, False otherwise
        """
        try:
            import csv

            if not data:
                return False

            headers = list(data[0].keys())

            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)

            logger.info(f"CSV file exported: {output_path}")
            return True

        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return False

    def export_transactions(self, transactions: List[Dict[str, Any]],
                           output_path: str) -> bool:
        """Export transactions to Excel."""
        formatted = []
        for t in transactions:
            formatted.append({
                "Date": t.get("date", ""),
                "Type": t.get("transaction_type", ""),
                "Category": t.get("category", ""),
                "Description": t.get("description", ""),
                "Amount": t.get("amount", 0),
                "Payment Method": t.get("payment_method", ""),
            })
        return self.export_to_excel(formatted, output_path, "Transactions")

    def export_budget_report(self, summary: Dict[str, Any],
                            output_path: str) -> bool:
        """Export budget summary to Excel."""
        formatted = [
            {"Metric": "Total Income", "Value": summary.get("total_income", 0)},
            {"Metric": "Total Expenses", "Value": summary.get("total_expenses", 0)},
            {"Metric": "Savings", "Value": summary.get("savings", 0)},
            {"Metric": "Savings Rate (%)", "Value": summary.get("savings_rate", 0)},
        ]

        # Add category breakdown
        for cat, amount in summary.get("category_breakdown", {}).items():
            formatted.append({"Metric": f"Expense: {cat}", "Value": amount})

        return self.export_to_excel(formatted, output_path, "Budget Report")


# Singleton instance
_excel_exporter = None


def get_excel_exporter() -> ExcelExporter:
    """Get or create the Excel exporter singleton."""
    global _excel_exporter
    if _excel_exporter is None:
        _excel_exporter = ExcelExporter()
    return _excel_exporter
